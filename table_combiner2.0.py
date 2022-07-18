#A quick and dirty script for organizing bondtesting data from Hatvan
#into a single .xlsx file, calculating failure mode percentages,
#creating individual value plots with force values and bar charts with
#failure mode percentages. An option for quickly grouping the tables is
#also included.

#No groups -> Plots all data into a single scatter plot/bar chart.

#Ready -> Opens second window with text boxes for each line entered.
#Each file is listed and numbered. The numbers may be entered in the group text
#boxes. Pressing the ready button plots the data by groups. A second sheet
#with the groups separated is also created in the combined .xlsx file.

import xlrd
import os
import re
import PyQt5.QtWidgets as qtw
import openpyxl
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib


def main():
    xls_regex = re.compile(r'\s*\.xls\Z')
    ero_regex = re.compile(r'\d+\.\d+')
    fmode_szam_regex = re.compile(r'\d{1}')
    wb = openpyxl.Workbook()
    ws = wb.active

    def minitab(ws, column_number):
        font = {
            'weight' : 'bold',
            'size'   : 15}
        matplotlib.rc('font', **font)
        row = 2
        column = 2
        group_column = 1
        df = pd.DataFrame(columns=['','Force [N]'])
        df2 = pd.DataFrame(columns = ['', 'Failure mode', 'Failure percentage'])
        while True:
            if ws.cell(row,column).value != None:
                force = float(ws.cell(row,column).value.replace(',','.'))
                group = ws.cell(1,group_column).value
                if group == 1:
                    group = ''
                df = df.append({'':group,'Force [N]':force}, ignore_index=True)
                row+=1
            elif ws.cell(2, column+10).value != None:
                row = 2
                column+=10
                group_column+=10
                continue
            else:
                break
        row = 1
        column = column_number
        group_column = 1
        while True:
            if ws.cell(row,column).value != None:
                fmode_percentage = float(ws.cell(row,column).value)
                fmode = ws.cell(row,column-1).value
                group = ws.cell(1, group_column).value
                if group == 1:
                    group = ''
                df2 = df2.append({'':group, 'Failure mode':fmode, 'Failure percentage':fmode_percentage}, ignore_index = True)
                row+=1
            elif ws.cell(1, column+10).value != None:
                row = 1
                column+=10
                group_column+=10
                continue
            else:
                break
        fig, axs = plt.subplots(2)
        fig.suptitle(minitab_neve)
        df = df.sort_values(by='')
        ax = sns.stripplot(x='',y='Force [N]', data=df, ax = axs[0])
        ax.set_ylim(0)
        df2 = df2.pivot(index = '',columns='Failure mode', values = 'Failure percentage')
        df2.plot.bar( rot=0, stacked=True, ax = axs[1])
        plt.ylim(0,100)
        plt.ylabel("Percentage [%]")
        ax.grid(axis='y')
        plt.show()
    
    mappa = input("\nMappa path copy, 1. sor:\n")
    neveben = input("\nFájlok nevében valami? (pl.: pull, shear, E2022401) \n.xls-t nem kell írni \nKihagyható, enterrel:\n")
    tabla_neve=input("\nEgyberakott tábla neve (kiterjesztés nélkül):\n")
    minitab_neve = input("\n'Minitab' diagram címe:\n")


    n = 1
    print(f"\"{tabla_neve}.xlsx\"-be pakolva:\n")
    
    
    fmode_dict = {}
    erok = []
    def get_fmode_oszlop():
        global fmode_szam_oszlop
        global fmode_oszlop
        global szorzo
        for root, dirs, files in os.walk(mappa, topdown = False):
            for fajl in files:
                if xls_regex.search(fajl) and neveben in fajl:
                    hatvani_tabla_path = os.path.join(root, fajl)
                    workbook = xlrd.open_workbook(hatvani_tabla_path)
                    sheet = workbook.sheet_by_index(0)
                    for row in range(0, sheet.nrows):
                        for column in range(0,sheet.ncols):
                            if "Resultcode (N" in str(sheet.cell_value(row, column)):
                                fmode_szam_oszlop = column
                                fmode_oszlop = fmode_szam_oszlop + 2
                            if sheet.cell_value(row, column) == "Peak force":
                                for column in range(0, sheet.ncols):
                                    if sheet.cell_value(row+1,column)=="cN":
                                        szorzo = 0.01
                                        return
                                    elif sheet.cell_value(row+1, column)=="gf":
                                        szorzo = 0.00980665
                                        return
                            
                                                   
    get_fmode_oszlop()
    for root, dirs, files in os.walk(mappa, topdown = False):
        for fajl in files:
            fmode_sor_helyi = n
            fmode_helyi_counter = 0
            fmode_dict_helyi = {}
    
            if xls_regex.search(fajl) and neveben in fajl:
                sorszam=1
                hatvani_tabla = fajl
                hatvani_tabla_path = os.path.join(root, fajl)
                ws.cell(n,4,hatvani_tabla)
                workbook = xlrd.open_workbook(hatvani_tabla_path)
                sheet = workbook.sheet_by_index(0)
                print(hatvani_tabla)
            
                for sor in range (0, sheet.nrows):
                    
                    ero = str(sheet.cell_value(sor, 3))
                    if ero_regex.search(ero):
                        ero_N = float(ero)*szorzo
                        erok.append(ero_N)
                        ero_N_str = str(ero_N)
                        ws.cell(n,1,sorszam)
                        ws.cell(n,2,ero_N_str.replace(".",","))
                        sorszam +=1
    
                    fmode_szam = str(sheet.cell_value(sor, fmode_szam_oszlop))
                    if fmode_szam_regex.search(fmode_szam):
                        fmode = sheet.cell_value(sor, fmode_oszlop)
                        ws.cell(n,3,fmode)
                        n +=1
                        if fmode not in fmode_dict:
                            fmode_dict[fmode] = 1
                        else:
                            fmode_dict[fmode] +=1
                            
                        if fmode not in fmode_dict_helyi:
                            fmode_dict_helyi[fmode] = 1
                            fmode_helyi_counter +=1
                        else:
                            fmode_dict_helyi[fmode] +=1
                            fmode_helyi_counter +=1
                            
                for fmode,fmode_mennyiseg in fmode_dict_helyi.items():
                    ws.cell(fmode_sor_helyi,5,fmode)
                    ws.cell(fmode_sor_helyi,6,fmode_mennyiseg/fmode_helyi_counter*100)
                    fmode_sor_helyi += 1

    fmode_sor = 1
    for fmode,fmode_mennyiseg in fmode_dict.items():
      ws.cell(fmode_sor,7,fmode)
      ws.cell(fmode_sor,8,fmode_mennyiseg/ws.max_row*100)
      fmode_sor += 1
      
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                  dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
        
    file_list = []
    file_list_str = ""
    
    file_dct = {}
    
    row = 1
    for i in range(ws.max_row):
        file_cell = ws.cell(row=row, column = 4)
        force= ws.cell(row=row, column = 2).value
        number = ws.cell(row=row,column = 1).value
        fmode = ws.cell(row=row, column = 3).value
        if file_cell.value != None:
            file = file_cell.value
            file_dct[file] = [(number,force,fmode,file)] 
            file_list.append(file_cell.value)
        else:
            file_dct[file].append((number,force,fmode,""))
        row +=1

    for idx, file in enumerate(file_list):
        file_list_str += f'{idx}-{file}\n'
        
    csoport_lista=[]
    
    class SecondWindow(qtw.QWidget):
        
        def __init__(self):
            super().__init__()
            layout = qtw.QVBoxLayout()
            self.setLayout(layout)
            self.boxes_labels()
            self.setWindowTitle("Enter group members")
            
        def boxes_labels(self):
            container= qtw.QWidget()
            container.setLayout(qtw.QGridLayout())
            self.boxes = []
            self.labels = []
            files_label = qtw.QLabel(file_list_str)
            container.layout().addWidget(files_label, 0,len(csoport_lista)+1,6,1)
            for idx, csoport in enumerate(csoport_lista):
                label = qtw.QLabel(csoport)
                self.labels.append(label)
                container.layout().addWidget(label, 0,idx,1,1)
                box = qtw.QPlainTextEdit()
                self.boxes.append(box)
                container.layout().addWidget(box, 1,idx,5,1)
            btn = qtw.QPushButton('Ready', clicked = lambda : self.grouper())
            container.layout().addWidget(btn,6,0,1,len(csoport_lista))
            self.layout().addWidget(container)
            
        def grouper(self):
            column = 1
            rendezett_sheet = wb.create_sheet(title = 'Rendezett')
            for idx, label in enumerate(self.labels):
                row = 1
                rendezett_sheet.cell(row, column, label.text())
                row+=1
                fmode_counter = 0
                fmode_dict2 = {}
                for i in self.boxes[idx].toPlainText().split('\n'):
                    file = file_list[int(i)]
                    for file_values in file_dct[file]:
                        rendezett_sheet.cell(row,column,file_values[0])
                        rendezett_sheet.cell(row,column+1,file_values[1])
                        rendezett_sheet.cell(row,column+2,file_values[2])
                        rendezett_sheet.cell(row,column+3,file_values[3])
                        if file_values[2] not in fmode_dict2:
                            fmode_dict2[file_values[2]] = 1
                            fmode_counter +=1
                        else:
                            fmode_dict2[file_values[2]] += 1
                            fmode_counter +=1
                        row+=1
                    fmode_sor = 1
                for fmode,fmode_mennyiseg in fmode_dict2.items():
                    rendezett_sheet.cell(fmode_sor,column+4,fmode)
                    rendezett_sheet.cell(fmode_sor,column+5,fmode_mennyiseg/fmode_counter*100)
                    fmode_sor += 1
                column+=10
            dims = {}
            for row in rendezett_sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                rendezett_sheet.column_dimensions[col].width = value
            wb.save(f"{tabla_neve}.xlsx")
            os.startfile(f"{tabla_neve}.xlsx")
            minitab(wb["Rendezett"],6)
            return
        
    class MainWindow(qtw.QWidget):
        
        def __init__(self):
            super().__init__()
            layout = qtw.QVBoxLayout()
            self.setLayout(layout)
            self.box_btn()
            self.setWindowTitle("Enter groups")
            
        def box_btn(self):
            container= qtw.QWidget()
            container.setLayout(qtw.QGridLayout())
            btn = qtw.QPushButton("Ready", clicked = lambda : self.groups())
            container.layout().addWidget(btn)
            btn2 = qtw.QPushButton("No groups", clicked = lambda : self.no_groups())
            container.layout().addWidget(btn2)
            self.box = qtw.QPlainTextEdit()
            container.layout().addWidget(self.box)
            self.layout().addWidget(container)
            
        def groups(self):
            for i in self.box.toPlainText().split('\n'):
                csoport_lista.append(i)
            self.sw = SecondWindow()
            self.sw.show()
            
        def no_groups(self):
            wb.save(f"{tabla_neve}.xlsx")
            os.startfile(f"{tabla_neve}.xlsx")
            minitab(wb.active, 8)
            return

    mw = MainWindow()
    mw.show()
    mw.raise_()

main()





