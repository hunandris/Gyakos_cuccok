import csv
import os
import re
import openpyxl

csv_regex = re.compile(r'\s*\.csv\Z')
ero_regex = re.compile(r'\d+\.\d+')
fmode_szam_regex = re.compile(r'\d{1}')
wb = openpyxl.Workbook()
ws = wb.active


def main():
    
    mappa = input("\nMappa path copy, 1. sor:\n")
    neveben = input("\nFájlok nevében valami? (pl.: pull, shear, E2022401) \n.csv-t nem kell írni \nKihagyható, enterrel:\n")
    tabla_neve=input("\nEgyberakott tábla neve (kiterjesztés nélkül):\n")

    
    n = 1
    print(f"\"{tabla_neve}.xlsx\"-be pakolva:\n")
    
    
    fmode_dict = {}
    erok = []
    for root, dirs, files in os.walk(mappa, topdown = False):
        for fajl in files:
            fmode_sor_helyi = n
            fmode_helyi_counter = 0
            fmode_dict_helyi = {}
        
            if csv_regex.search(fajl) and neveben in fajl:
                sorszam=1
                sajat_tabla = fajl
                ws.cell(n,5,sajat_tabla)
                path = os.path.join(root,fajl)
                print(sajat_tabla)
                with open (path, newline = '') as csvfile:
                    csvreader = csv.reader(csvfile, delimiter = ';' )

       
                    
                    for index, row in enumerate(csvreader):
                        for column in range(len(row)):
                            if row[column] == "Force":
                                starting_row = index+1
                                ero_column = column
                            elif row[column] == "Type or grade":
                                fmode_column = column
                            elif row[column] == "Comment":
                                comment_column = column
                            else:
                                continue
                    
                    csvfile.seek(0)
                    csvreader.__init__(csvfile, delimiter = ';' )
                    for i, r in enumerate(csvreader):
                        if i == starting_row:
                            try:
                                ero = r[ero_column]
                                erok.append(ero)
                                ws.cell(n,1, sorszam)
                                ws.cell(n,2,ero)
                                sorszam += 1
                                
                                fmode_szam = r[fmode_column]
                                ws.cell(n,3,fmode_szam)
                                
                                
                                comment = r[comment_column]
                                ws.cell(n,4, comment)
                                
                                n +=1
                                    
                                if fmode_szam not in fmode_dict:
                                    fmode_dict[fmode_szam] = 1
                                else:
                                    fmode_dict[fmode_szam] +=1
                                    
                                if fmode_szam not in fmode_dict_helyi:
                                    fmode_dict_helyi[fmode_szam] = 1
                                    fmode_helyi_counter +=1
                                else:
                                    fmode_dict_helyi[fmode_szam] +=1
                                    fmode_helyi_counter +=1
                            except IndexError:
                                 break
                                
                        else:
                            continue
                        starting_row += 1
                                    
            for fmode_szam,fmode_mennyiseg in fmode_dict_helyi.items():
                ws.cell(fmode_sor_helyi,6,fmode_szam)
                ws.cell(fmode_sor_helyi,7,fmode_mennyiseg/fmode_helyi_counter*100)
                fmode_sor_helyi += 1                  
              

    fmode_sor = 1
    for fmode_szam,fmode_mennyiseg in fmode_dict.items():
      ws.cell(fmode_sor,8,fmode_szam)
      ws.cell(fmode_sor,9,fmode_mennyiseg/ws.max_row*100)
      fmode_sor += 1
      
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                  dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
        
    file_list = []
    file_list_str = ""
    
    file_dct = {}
    
    row = 1
    for i in range(ws.max_row):
        file_cell = ws.cell(row=row, column = 5)
        force= ws.cell(row=row, column = 2).value
        number = ws.cell(row=row,column = 1).value
        fmode = ws.cell(row=row, column = 3).value
        if file_cell.value != None:
            file = file_cell.value
            file_dct[file] = [(number,force,fmode,file)] 
            file_list.append(file_cell.value)
        else:
            file_dct[file].append((number,force,fmode,""))
        row +=1

    for idx, file in enumerate(file_list):
        file_list_str += f'{idx}-{file}\n'
        
    wb.save(f"{tabla_neve}.xlsx")
    os.startfile(f"{tabla_neve}.xlsx")
main()
   





